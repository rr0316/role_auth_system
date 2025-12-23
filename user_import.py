import openpyxl
import os
from datetime import datetime
from database import Database
from auth_system import AuthSystem

class UserImport:
    """Excel批量导入用户"""
    def __init__(self):
        self.db = Database()
        self.auth = AuthSystem()
        self.import_report = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'duplicate': 0,
            'failed_records': [],
            'duplicate_records': []
        }

    def _read_excel(self, file_path):
        """读取Excel文件"""
        if not os.path.exists(file_path):
            raise Exception(f"文件不存在：{file_path}")
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 验证表头
        headers = [cell.value for cell in ws[1]]
        required_headers = ['用户编码', '角色', '真实姓名', '密码', '手机号', '邮箱']
        if headers[:6] != required_headers:
            raise Exception(f"Excel表头错误，必须为：{required_headers}")
        
        # 读取数据（跳过表头）
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # 空行跳过
                continue
            data.append({
                'user_code': str(row[0]).strip() if row[0] else None,
                'role': row[1].strip().lower() if row[1] else None,
                'real_name': row[2].strip() if row[2] else None,
                'password': row[3].strip() if row[3] else None,
                'phone': row[4].strip() if row[4] else None,
                'email': row[5].strip() if row[5] else None,
                'row_num': ws._current_row  # 行号
            })
        wb.close()
        return data

    def import_users(self, file_path, default_password='123456'):
        """
        批量导入用户
        :param file_path: Excel文件路径
        :param default_password: 默认密码
        :return: 导入报告
        """
        try:
            # 1. 读取Excel数据
            users = self._read_excel(file_path)
            self.import_report['total'] = len(users)
            
            # 2. 逐条验证并导入
            for user in users:
                user_code = user['user_code']
                role = user['role']
                real_name = user['real_name']
                password = user['password'] or default_password
                
                # 基础验证
                if not all([user_code, role, real_name]):
                    self.import_report['failed'] += 1
                    self.import_report['failed_records'].append({
                        'row': user['row_num'],
                        'user_code': user_code,
                        'reason': '用户编码/角色/真实姓名不能为空'
                    })
                    continue
                
                # 角色验证
                if role not in ['student', 'teacher', 'admin']:
                    self.import_report['failed'] += 1
                    self.import_report['failed_records'].append({
                        'row': user['row_num'],
                        'user_code': user_code,
                        'reason': f'角色错误（{role}），必须为student/teacher/admin'
                    })
                    continue
                
                # 检查重复
                if self.db.check_user_code_exists(user_code):
                    self.import_report['duplicate'] += 1
                    self.import_report['duplicate_records'].append({
                        'row': user['row_num'],
                        'user_code': user_code,
                        'reason': '账号已存在'
                    })
                    continue
                
                # 格式验证
                valid, msg = self.auth._validate_user_code(user_code, role)
                if not valid:
                    self.import_report['failed'] += 1
                    self.import_report['failed_records'].append({
                        'row': user['row_num'],
                        'user_code': user_code,
                        'reason': msg
                    })
                    continue
                
                # 注册用户
                success, msg = self.auth.register(
                    user_code=user_code,
                    password=password,
                    role=role,
                    real_name=real_name,
                    phone=user['phone'],
                    email=user['email']
                )
                if success:
                    self.import_report['success'] += 1
                else:
                    self.import_report['failed'] += 1
                    self.import_report['failed_records'].append({
                        'row': user['row_num'],
                        'user_code': user_code,
                        'reason': msg
                    })
            
            # 3. 生成导入报告
            self._generate_report()
            return self.import_report
        
        except Exception as e:
            raise Exception(f"导入失败：{str(e)}")

    def _generate_report(self):
        """生成导入测试报告（Markdown格式）"""
        report_content = f"""# 用户批量导入测试报告
## 导入基本信息
- 导入时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- 总记录数：{self.import_report['total']}
- 成功数：{self.import_report['success']}
- 失败数：{self.import_report['failed']}
- 重复数：{self.import_report['duplicate']}

## 失败记录详情
| 行号 | 用户编码 | 失败原因 |
|------|----------|----------|
"""
        for record in self.import_report['failed_records']:
            report_content += f"| {record['row']} | {record['user_code'] or '空'} | {record['reason']} |\n"
        
        report_content += """
## 重复记录详情
| 行号 | 用户编码 | 重复原因 |
|------|----------|----------|
"""
        for record in self.import_report['duplicate_records']:
            report_content += f"| {record['row']} | {record['user_code'] or '空'} | {record['reason']} |\n"
        
        # 保存报告
        with open('import_test_report.md', 'w', encoding='utf-8') as f:
            f.write(report_content)

# 测试示例
if __name__ == "__main__":
    importer = UserImport()
    try:
        # 导入Excel文件（需提前创建sample_users.xlsx）
        report = importer.import_users('sample_users.xlsx')
        print("导入报告：", report)
        print("报告已生成：import_test_report.md")
    except Exception as e:
        print("导入失败：", e)
